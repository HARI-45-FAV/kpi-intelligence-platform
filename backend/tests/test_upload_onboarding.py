"""CSV and Excel onboarding: what the platform infers, and what it admits to.

An upload is the one place where the platform decides, on its own, what somebody
else's data *means* — that a column of digits is a measure rather than a code, that
``01/02`` is the 1st of February. Every one of those decisions ends up underneath a
KPI, so this module is mostly about honesty rather than about parsing:

1. **Preview writes nothing.** A reader sees the inferred types and every assumption
   before a source exists, and refusing costs them nothing.
2. **Inference is conservative and disclosed.** Leading-zero codes stay text,
   ambiguous dates say they were ambiguous, symbol-stripping says it stripped.
3. **Numbers that are not loaded are not counted.** A malformed row is skipped,
   counted, and reported separately from the rows that landed.
4. **An upload becomes an ordinary SQL source.** Discovery, selection and profiling
   are the platform's existing ones, unchanged — which is the whole reason the file
   is loaded into a database rather than held in memory.
5. **The record outlives the rows.** Every load is in the audit trail with its
   filename, checksum, row counts and assumptions, so "which export is this KPI
   built on" survives the next re-upload.
6. **The boundaries hold.** ``source.manage`` is required, another company cannot
   reach the source, a filename cannot escape the storage directory, and an
   identical file is refused rather than reloaded.

Everything is driven through the same HTTP API the frontend uses. Two reads go to
the ORM instead: the stored ``storage_path``, which is deliberately never exposed
over the API, and the loaded rows themselves, which are asserted with the stdlib
driver so "the file was replaced, not appended to" is proved against the data and
not against a number the loader reported about itself.
"""

from __future__ import annotations

import hashlib
import io
import sqlite3
from collections.abc import Iterator
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.core.config import settings
from app.core.database import SessionLocal
from app.main import create_app
from app.models.base import DataSourceType, RefreshFrequency
from app.models.source import DataSource, SourceTable
from tests.conftest import API, ApiActor, login, register

PASSWORD = "Upload-Onboarding-2026"

# A deliberately awkward export. Every column is here to force one decision:
#
#   order_date    ISO dates, unambiguous          -> DATE, no assumption needed
#   region        plain text                      -> TEXT
#   product_code  leading zeros                   -> TEXT, not INTEGER
#   units         plain integers                  -> INTEGER
#   net_revenue   currency, thousands, ( ) losses -> REAL, and says what it removed
#   active        true/false words                -> BOOLEAN
#   notes         one value, one 'n/a', two empty -> TEXT, missing counted honestly
#
# The last line has three fields against a seven-field header: a real export
# truncated mid-write. It must be skipped and counted, never padded.
MESSY_CSV = (
    "order_date,region,product_code,units,net_revenue,active,notes\n"
    '2026-08-01,North,007,3,"$1,250.50",true,\n'
    "2026-08-02,South,012,5,$980.00,false,late\n"
    "2026-08-03,North,007,2,($120.25),true,n/a\n"
    '2026-08-04,East,900,7,"$3,400.75",TRUE,\n'
    "2026-08-05,West,011\n"
).encode()

# The same shape, one row shorter and with a different region, for the reload tests.
REVISED_CSV = (
    "order_date,region,product_code,units,net_revenue,active,notes\n"
    '2026-08-01,North,007,3,"$1,250.50",true,\n'
    "2026-08-02,South,012,5,$980.00,false,late\n"
    '2026-08-06,Central,015,9,"$5,000.00",true,new\n'
).encode()

# Digits that parse under both day-first and month-first readings, and under nothing
# else. There is no way to tell them apart from the file, which is the point.
AMBIGUOUS_CSV = (
    "txn_date,amount\n01/02/2026,10\n03/04/2026,20\n05/06/2026,30\n"
).encode()


def _csv_upload(name: str, data: bytes) -> dict:
    return {"file": (name, io.BytesIO(data), "text/csv")}


def _workbook() -> bytes:
    """A two-sheet workbook whose cells carry real Excel types.

    Built with openpyxl rather than checked in as a binary so the expected types are
    visible here: ``booked_on`` is a genuine date cell, ``amount`` a float, ``qty`` an
    int. Nothing is a string, so the parser has no text to re-parse and should trust
    what the file already states.
    """

    from datetime import date

    from openpyxl import Workbook

    book = Workbook()
    sales = book.active
    sales.title = "Sales"
    sales.append(["booked_on", "channel", "qty", "amount"])
    sales.append([date(2026, 8, 1), "Retail", 3, 1250.5])
    sales.append([date(2026, 8, 2), "Online", 5, 980.0])
    sales.append([date(2026, 8, 3), "Retail", 2, 120.25])

    regions = book.create_sheet("Region Master")
    regions.append(["Region Code", "Region Name"])
    regions.append(["N", "North"])
    regions.append(["S", "South"])

    # An empty sheet is normal in a real workbook and must not fail the upload.
    book.create_sheet("Notes")

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Provisioning
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module", autouse=True)
def _upload_storage(tmp_path_factory) -> Iterator[Path]:
    """Uploads land in a throwaway directory for the length of this module.

    Patched on the settings object rather than the environment because
    ``_storage_path`` reads it per call, and because the tests assert that the file
    written is *inside* this directory — an assertion that means nothing if the
    directory is the developer's real one.
    """

    directory = tmp_path_factory.mktemp("uploads")
    original = settings.upload_storage_dir
    settings.upload_storage_dir = str(directory)
    yield directory
    settings.upload_storage_dir = original


@pytest.fixture(scope="module")
def module_client() -> Iterator[TestClient]:
    app = create_app()
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture(scope="module")
def workspace(module_client: TestClient) -> dict:
    admin = register(
        module_client, "uma@upload-co.example.com", PASSWORD, "Uma Upload"
    )
    created = admin.post(
        f"{API}/companies",
        json={"company_name": "Upload Co", "currency": "INR", "timezone": "Asia/Kolkata"},
    )
    assert created.status_code == 201, created.text
    company_id = created.json()["id"]
    base = f"{API}/companies/{company_id}"

    analyst = module_client.post(
        f"{base}/members",
        json={
            "email": "ravi@upload-co.example.com",
            "full_name": "Ravi Analyst",
            "password": PASSWORD,
            "role_key": "ANALYST",
        },
        headers=admin.headers,
    )
    assert analyst.status_code == 201, analyst.text

    return {
        "client": module_client,
        "admin": admin,
        "company_id": company_id,
        "base": base,
    }


def _columns(preview: dict, table_index: int = 0) -> dict[str, dict]:
    return {
        column["name"]: column
        for column in preview["tables"][table_index]["columns"]
    }


def _upload(workspace: dict, name: str, filename: str, data: bytes) -> dict:
    response = workspace["admin"].post(
        f"{workspace['base']}/uploads",
        files=_csv_upload(filename, data),
        data={"name": name, "profile": "true"},
    )
    assert response.status_code == 201, response.text
    return response.json()


# ---------------------------------------------------------------------------
# 1. Preview
# ---------------------------------------------------------------------------
def test_preview_reports_what_it_would_create_and_creates_nothing(workspace):
    """The consent step: every inference visible, no source registered.

    Asserted together because either half alone is the wrong feature. A preview
    that creates a source is not a preview, and a preview that hides its
    assumptions is not consent.
    """

    admin, base = workspace["admin"], workspace["base"]
    before = admin.get(f"{base}/data-sources").json()

    response = admin.post(
        f"{base}/uploads/preview", files=_csv_upload("august_orders.csv", MESSY_CSV)
    )
    assert response.status_code == 200, response.text
    preview = response.json()["preview"]

    assert preview["filename"] == "august_orders.csv"
    assert preview["file_format"] == "Delimited text"
    assert preview["checksum_sha256"] == hashlib.sha256(MESSY_CSV).hexdigest()
    assert len(preview["tables"]) == 1

    table = preview["tables"][0]
    assert table["table_name"] == "august_orders"
    assert table["row_count"] == 4
    # The malformed row is reported as its own number rather than folded into the
    # row count: "4 rows" and "5 lines in the file" are both true and a reader
    # reconciling the upload against their export needs to see both.
    assert table["skipped_rows"] == 1
    assert any("different number of fields" in note for note in preview["notes"])

    # Nothing was written.
    assert admin.get(f"{base}/data-sources").json() == before


def test_inference_is_conservative_and_says_what_it_assumed(workspace):
    """Types, and the assumption behind each one that needed a decision."""

    admin, base = workspace["admin"], workspace["base"]
    preview = admin.post(
        f"{base}/uploads/preview", files=_csv_upload("august_orders.csv", MESSY_CSV)
    ).json()["preview"]
    columns = _columns(preview)

    assert columns["order_date"]["sql_type"] == "DATE"
    assert columns["order_date"]["type_family"] == "TEMPORAL"
    assert columns["region"]["sql_type"] == "TEXT"
    assert columns["units"]["sql_type"] == "INTEGER"
    assert columns["active"]["sql_type"] == "BOOLEAN"

    # A leading zero is significant in a code: '007' is a store, not seven. Read as
    # a number it would come back as 7 and never match the source system again.
    assert columns["product_code"]["sql_type"] == "TEXT"
    assert columns["product_code"]["distinct_sample"][0] == "007"

    # Currency, thousands separators and a parenthesised loss are formatting, not
    # data — but removing them is still an assumption, so it is stated.
    revenue = columns["net_revenue"]
    assert revenue["sql_type"] == "REAL"
    assert revenue["type_family"] == "NUMERIC"
    assert any("currency" in note.lower() for note in revenue["notes"])
    assert revenue["distinct_sample"][:3] == [1250.5, 980.0, -120.25]

    # 'n/a' is missing, not the three-character string 'n/a' — and the two are
    # counted apart, because "this cell was empty" and "this cell said n/a" are
    # different facts about the export.
    notes_column = columns["notes"]
    assert notes_column["filled"] == 1
    assert notes_column["blank"] == 3
    assert notes_column["blank_tokens"] == 1
    assert any("'n/a'" in note for note in notes_column["notes"])


def test_an_ambiguous_date_says_which_way_round_it_was_read(workspace):
    """01/02/2026 is two different days. The reader is told which one was chosen."""

    admin, base = workspace["admin"], workspace["base"]
    preview = admin.post(
        f"{base}/uploads/preview", files=_csv_upload("transactions.csv", AMBIGUOUS_CSV)
    ).json()["preview"]
    column = _columns(preview)["txn_date"]

    assert column["sql_type"] == "DATE"
    assert any("either way round" in note for note in column["notes"])
    # Day-before-month, as the note says: the 1st of February, not the 2nd of January.
    assert column["distinct_sample"][0] == "2026-02-01"


def test_a_table_with_no_date_column_says_so(workspace):
    """Not an error — but it decides whether a KPI over time can be built here."""

    admin, base = workspace["admin"], workspace["base"]
    preview = admin.post(
        f"{base}/uploads/preview",
        files=_csv_upload("regions.csv", b"region_code,region_name\nN,North\nS,South\n"),
    ).json()["preview"]

    assert any("No date column" in note for note in preview["tables"][0]["notes"])


def test_an_unreadable_format_is_refused_with_the_way_out(workspace):
    """A refusal that does not say what to do instead is a dead end."""

    admin, base = workspace["admin"], workspace["base"]
    refused = admin.post(
        f"{base}/uploads/preview", files=_csv_upload("legacy.xls", b"\xd0\xcf\x11\xe0junk")
    )
    assert refused.status_code == 422
    message = refused.json()["message"]
    assert ".xlsx" in message and ".csv" in message

    empty = admin.post(f"{base}/uploads/preview", files=_csv_upload("empty.csv", b""))
    assert empty.status_code == 422
    assert "empty" in empty.json()["message"].lower()


# ---------------------------------------------------------------------------
# 2. The upload becomes an ordinary source
# ---------------------------------------------------------------------------
def test_an_uploaded_file_is_discovered_and_profiled_like_any_source(workspace):
    """The payoff for loading into SQL: nothing downstream is re-implemented.

    Discovery reflects the created table, selection enables it, and the platform's
    own profiler measures it — so an uploaded sheet arrives in the catalog with the
    same metadata as a table in a warehouse. The date column reaching
    ``semantic_type: DATE`` is the load-bearing assertion: it is what makes the
    column eligible as a KPI time field, and it only happens because the loader
    declared DATE rather than TEXT.
    """

    admin, base = workspace["admin"], workspace["base"]
    result = _upload(workspace, "August Orders", "august_orders.csv", MESSY_CSV)

    assert result["load"]["total_rows"] == 4
    discovery = result["discovery"]
    assert discovery["tables_found"] == 1
    assert discovery["tables_created"] == 1
    assert discovery["columns_created"] == 7
    assert discovery["tables_selected"] == 1
    assert discovery["tables_profiled"] == ["august_orders"]
    assert discovery["profiling_failures"] == []

    source = admin.get(f"{base}/data-sources/{result['source_id']}").json()
    assert source["source_type"] == DataSourceType.UPLOAD
    assert source["refresh_frequency"] == RefreshFrequency.MANUAL
    assert source["connection_status"] == "CONNECTED"

    tables = {row["table_name"]: row for row in admin.get(f"{base}/tables").json()}
    assert "august_orders" in tables
    table_id = tables["august_orders"]["id"]

    profile = admin.get(f"{base}/tables/{table_id}/profile").json()
    assert profile["profile"] is not None
    assert profile["profile"]["row_count"] == 4

    columns = {
        column["column_name"]: column
        for column in admin.get(f"{base}/tables/{table_id}/columns").json()
    }
    assert columns["order_date"]["semantic_type"] == "DATE"
    # The declared storage type is what the loader controls and what reflection
    # reports, so that is what is asserted here. The *semantic* type is the
    # profiler's own reading and depends on cardinality — on four rows it may call a
    # small integer categorical, which is a reasonable thing for it to say about four
    # rows and not something this feature should pin.
    assert "INT" in columns["units"]["data_type"].upper()
    assert "REAL" in columns["net_revenue"]["data_type"].upper()
    assert "DATE" in columns["order_date"]["data_type"].upper()
    assert columns["product_code"]["data_type"].upper() in ("TEXT", "VARCHAR")


def test_the_loaded_rows_are_the_converted_ones(workspace):
    """The data on disk, read with the stdlib driver rather than trusted.

    Asserted against the file because everything else in the platform reads this
    table through the connector, so if the conversion were wrong every downstream
    number would be wrong in the same direction and agree with itself.
    """

    result = _upload(workspace, "Rows On Disk", "august_orders.csv", MESSY_CSV)

    with SessionLocal() as session:
        source = session.get(DataSource, result["source_id"])
        path = Path(source.options["storage_path"])

    assert path.exists()
    # Inside the configured directory, and named after ids the platform issued —
    # never after anything the caller sent.
    assert path.is_relative_to(Path(settings.upload_storage_dir))
    assert path.name == f"{result['source_id']}.db"

    connection = sqlite3.connect(path)
    try:
        rows = connection.execute(
            "SELECT order_date, product_code, units, net_revenue, active "
            "FROM august_orders ORDER BY order_date"
        ).fetchall()
    finally:
        connection.close()

    assert rows == [
        ("2026-08-01", "007", 3, 1250.5, 1),
        ("2026-08-02", "012", 5, 980.0, 0),
        # A parenthesised figure is a loss, and it is stored as one.
        ("2026-08-03", "007", 2, -120.25, 1),
        ("2026-08-04", "900", 7, 3400.75, 1),
    ]


def test_a_workbook_becomes_one_table_per_sheet(workspace):
    """Excel's own cell types are trusted rather than re-derived from text."""

    admin, base = workspace["admin"], workspace["base"]
    response = admin.post(
        f"{base}/uploads",
        files={
            "file": (
                "trading.xlsx",
                io.BytesIO(_workbook()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"name": "Trading Workbook", "profile": "true"},
    )
    assert response.status_code == 201, response.text
    body = response.json()

    preview = body["preview"]
    assert preview["file_format"] == "Excel workbook"
    tables = {table["table_name"]: table for table in preview["tables"]}
    # Two sheets with data become two tables; the empty third is skipped and said so.
    assert set(tables) == {"sales", "region_master"}
    assert any("held no data rows" in note for note in preview["notes"])
    assert body["load"]["total_rows"] == 5

    sales = {column["name"]: column for column in tables["sales"]["columns"]}
    assert sales["booked_on"]["sql_type"] == "DATE"
    assert sales["qty"]["sql_type"] == "INTEGER"
    assert sales["amount"]["sql_type"] == "REAL"
    # No assumption was needed: the file already stated these types.
    assert sales["amount"]["notes"] == []

    master = {column["name"]: column for column in tables["region_master"]["columns"]}
    # A heading that is not a usable identifier is renamed, and the original kept.
    assert any("Region Code" in note for note in master["region_code"]["notes"])

    discovered = {row["table_name"] for row in admin.get(f"{base}/tables").json()}
    assert {"sales", "region_master"} <= discovered


# ---------------------------------------------------------------------------
# 3. The record, and the reload
# ---------------------------------------------------------------------------
def test_every_load_is_in_the_audit_trail_with_its_assumptions(workspace):
    """What the numbers were built on, in a place the browser tab cannot lose."""

    admin, base = workspace["admin"], workspace["base"]
    result = _upload(workspace, "Audited Orders", "august_orders.csv", MESSY_CSV)

    entries = admin.get(
        f"{base}/audit", params={"action": "source.file_uploaded", "limit": 50}
    ).json()
    mine = [entry for entry in entries if entry["resource_id"] == result["source_id"]]
    assert len(mine) == 1

    details = mine[0]["details"]
    assert details["filename"] == "august_orders.csv"
    assert details["checksum_sha256"] == hashlib.sha256(MESSY_CSV).hexdigest()
    assert details["rows_loaded"] == 4
    assert details["rows_skipped"] == 1
    assumptions = details["assumptions"]
    assert any("currency" in line.lower() for line in assumptions)
    assert any("skipped rather than padded" in line for line in assumptions)


def test_reloading_replaces_the_rows_and_names_what_disappeared(workspace):
    """A re-export is a new statement of the same facts, not more facts."""

    admin, base = workspace["admin"], workspace["base"]
    result = _upload(workspace, "Reloadable Orders", "august_orders.csv", MESSY_CSV)
    source_id = result["source_id"]

    reloaded = admin.post(
        f"{base}/uploads/{source_id}/reload",
        files=_csv_upload("august_orders.csv", REVISED_CSV),
        data={"profile": "true"},
    )
    assert reloaded.status_code == 200, reloaded.text
    body = reloaded.json()
    assert body["load"]["total_rows"] == 3
    assert body["tables_no_longer_present"] == []

    with SessionLocal() as session:
        source = session.get(DataSource, source_id)
        path = Path(source.options["storage_path"])
        assert source.options["loaded_rows"] == 3

    connection = sqlite3.connect(path)
    try:
        regions = [
            row[0]
            for row in connection.execute(
                "SELECT DISTINCT region FROM august_orders ORDER BY region"
            )
        ]
        count = connection.execute("SELECT COUNT(*) FROM august_orders").fetchone()[0]
    finally:
        connection.close()

    # Replaced, not appended: three rows, and the regions from the first export that
    # are absent from the second are gone rather than doubled up alongside it.
    assert count == 3
    assert regions == ["Central", "North", "South"]

    # Both loads are still in the trail, which is what the rows no longer say.
    entries = admin.get(
        f"{base}/audit", params={"action": "source.file_uploaded", "limit": 50}
    ).json()
    mine = [entry for entry in entries if entry["resource_id"] == source_id]
    assert len(mine) == 2
    checksums = {entry["details"]["checksum_sha256"] for entry in mine}
    assert checksums == {
        hashlib.sha256(MESSY_CSV).hexdigest(),
        hashlib.sha256(REVISED_CSV).hexdigest(),
    }


def test_a_table_that_vanished_between_exports_is_named(workspace):
    """The failure that silently breaks a KPI, said out loud."""

    admin, base = workspace["admin"], workspace["base"]
    response = admin.post(
        f"{base}/uploads",
        files={"file": ("book.xlsx", io.BytesIO(_workbook()), "application/vnd.ms-excel")},
        data={"name": "Shrinking Workbook", "profile": "false"},
    )
    assert response.status_code == 201, response.text
    source_id = response.json()["source_id"]

    # The next export dropped a sheet.
    reloaded = admin.post(
        f"{base}/uploads/{source_id}/reload",
        files=_csv_upload("book.csv", b"region_code,region_name\nN,North\n"),
        data={"profile": "false"},
    )
    assert reloaded.status_code == 200, reloaded.text
    assert reloaded.json()["tables_no_longer_present"] == ["region_master", "sales"]


def test_an_identical_file_is_refused_rather_than_reloaded(workspace):
    """Reloading unchanged rows would record that new data arrived. None did."""

    admin, base = workspace["admin"], workspace["base"]
    result = _upload(workspace, "Idempotent Orders", "august_orders.csv", MESSY_CSV)

    refused = admin.post(
        f"{base}/uploads/{result['source_id']}/reload",
        files=_csv_upload("august_orders_copy.csv", MESSY_CSV),
        data={"profile": "false"},
    )
    assert refused.status_code == 409
    assert "byte-for-byte" in refused.json()["message"]

    entries = admin.get(
        f"{base}/audit", params={"action": "source.file_uploaded", "limit": 50}
    ).json()
    mine = [entry for entry in entries if entry["resource_id"] == result["source_id"]]
    assert len(mine) == 1


def test_a_duplicate_source_name_is_refused_with_the_alternative(workspace):
    admin, base = workspace["admin"], workspace["base"]
    _upload(workspace, "Unique Orders", "august_orders.csv", MESSY_CSV)

    clash = admin.post(
        f"{base}/uploads",
        files=_csv_upload("august_orders.csv", REVISED_CSV),
        data={"name": "Unique Orders"},
    )
    assert clash.status_code == 409
    assert "Re-upload into it" in clash.json()["message"]


# ---------------------------------------------------------------------------
# 4. Boundaries
# ---------------------------------------------------------------------------
def test_a_filename_cannot_escape_the_storage_directory(workspace):
    """A browser can send anything as a filename. None of it reaches the disk."""

    admin, base = workspace["admin"], workspace["base"]
    response = admin.post(
        f"{base}/uploads",
        files=_csv_upload("../../../../etc/passwd.csv", MESSY_CSV),
        data={"name": "Traversal Attempt"},
    )
    assert response.status_code == 201, response.text
    source_id = response.json()["source_id"]

    with SessionLocal() as session:
        source = session.get(DataSource, source_id)
        # The name is kept for the trail, stripped of every path component.
        assert source.options["original_filename"] == "passwd.csv"
        path = Path(source.options["storage_path"])

    assert path.is_relative_to(Path(settings.upload_storage_dir))
    assert path.name == f"{source_id}.db"


def test_uploading_needs_source_manage(workspace):
    """The same permission as registering a database, because that is what it is."""

    base = workspace["base"]
    analyst = login(
        workspace["client"], "ravi@upload-co.example.com", PASSWORD, workspace["company_id"]
    )

    preview = analyst.post(
        f"{base}/uploads/preview", files=_csv_upload("orders.csv", MESSY_CSV)
    )
    assert preview.status_code == 403

    created = analyst.post(
        f"{base}/uploads",
        files=_csv_upload("orders.csv", MESSY_CSV),
        data={"name": "Analyst Upload"},
    )
    assert created.status_code == 403


def test_another_companys_upload_cannot_be_reached(workspace):
    """Company scoping is re-derived from the database, not read from the URL."""

    client = workspace["client"]
    result = _upload(workspace, "Private Orders", "august_orders.csv", MESSY_CSV)

    outsider = register(client, "zara@other-co.example.com", PASSWORD, "Zara Other")
    other = outsider.post(
        f"{API}/companies",
        json={"company_name": "Other Co", "currency": "USD", "timezone": "UTC"},
    )
    assert other.status_code == 201, other.text
    other_base = f"{API}/companies/{other.json()['id']}"

    # Aimed at their own company, naming a source that is not theirs.
    refused = outsider.post(
        f"{other_base}/uploads/{result['source_id']}/reload",
        files=_csv_upload("august_orders.csv", REVISED_CSV),
        data={"profile": "false"},
    )
    assert refused.status_code == 404

    # And aimed at the owning company, which they are not a member of.
    trespass = outsider.post(
        f"{workspace['base']}/uploads/{result['source_id']}/reload",
        files=_csv_upload("august_orders.csv", REVISED_CSV),
        data={"profile": "false"},
    )
    assert trespass.status_code in (403, 404)


def test_reload_refuses_a_source_that_is_not_an_upload(workspace, source_fixture):
    """Replacing the rows of a connected database is not a thing this endpoint does."""

    admin, base = workspace["admin"], workspace["base"]
    registered = admin.post(
        f"{base}/data-sources",
        json={
            "name": "Connected NovaMart",
            "source_type": "SQLITE",
            "path": source_fixture["path"],
            "refresh_frequency": "DAILY",
            "timezone": "Asia/Kolkata",
        },
    )
    assert registered.status_code == 201, registered.text

    refused = admin.post(
        f"{base}/uploads/{registered.json()['id']}/reload",
        files=_csv_upload("orders.csv", MESSY_CSV),
        data={"profile": "false"},
    )
    assert refused.status_code == 422
    assert "not an upload" in refused.json()["message"]


def test_the_upload_records_that_it_will_not_refresh_itself(workspace):
    """A snapshot that claims a cadence would be reported stale forever."""

    result = _upload(workspace, "Snapshot Orders", "august_orders.csv", MESSY_CSV)

    with SessionLocal() as session:
        source = session.get(DataSource, result["source_id"])
        assert source.refresh_frequency == RefreshFrequency.MANUAL
        assert "does not refresh on its own" in (source.known_limitations or "")
        assert source.options["skipped_rows"] == 1

        # One table registered per loaded sheet, under this source only.
        tables = list(
            session.scalars(
                select(SourceTable).where(SourceTable.data_source_id == source.id)
            )
        )
        assert [table.table_name for table in tables] == ["august_orders"]
