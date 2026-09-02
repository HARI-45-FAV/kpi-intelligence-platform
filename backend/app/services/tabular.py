"""Reading a spreadsheet into something the platform can already govern.

An uploaded CSV or Excel sheet is the one source a company can hand over without a
DBA, and it is where most KPI work actually starts. The tempting shortcut is a
parallel pipeline: read the file into memory, compute the KPI in Python, draw a
chart. That would leave the platform with two analytical engines and two sets of
bugs -- and the file-based one would be the engine without grain detection, join
safety, row scope or lineage.

So this module does something narrower. It parses the file, decides one type per
column, and writes the rows into a SQLite database owned by the company. From there
the upload is a SQL source like any other: ``discover_source`` reflects it,
``profile_table`` pushes aggregates down to it, grain and relationship detection run
over it unchanged, and a KPI built on it is evaluated by the same formula engine
that evaluates one built on PostgreSQL. Nothing downstream learns the source began
as a file.

Four things this module refuses to do, because each one would put invented data
under a governed KPI:

* **Guess a type from the first row.** Inference reads every value in the column
  and falls back to text the moment one disagrees. A column of 999 integers and one
  ``"n/a"`` is text, because storing it as a number means deciding what ``"n/a"``
  meant.
* **Repair a malformed row.** A row with the wrong number of fields is skipped and
  counted, and the count is returned to the caller and written to the audit trail.
  Padding a short row with nulls fabricates values.
* **Convert without saying so.** Stripping a currency symbol or reading
  ``31/01/2026`` as a date is a decision about someone's data. Each one is recorded
  as a note on the column, so the reader sees what was assumed before they build a
  KPI on it.
* **Preserve a header it cannot use as an identifier.** Names are normalised to
  satisfy ``connectors.base.validate_identifier``, and the original header is kept
  verbatim as the column comment, so the mapping stays inspectable rather than
  becoming folklore.
"""

from __future__ import annotations

import csv
import io
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from app.connectors.base import validate_identifier
from app.core.errors import ValidationFailure

# ---------------------------------------------------------------------------
# What counts as a blank
# ---------------------------------------------------------------------------
# Treating "n/a" as absent is an interpretation, and an unavoidable one: refusing it
# would type almost every real export as text. The compromise is a small, explicit
# set -- no regex, no "anything short and non-numeric" -- and a per-column count of
# how many values it matched, reported to the caller. The platform says what it did
# rather than quietly deciding.
_BLANK_TOKENS = frozenset(
    {"", "na", "n/a", "n.a.", "null", "none", "nil", "-", "--", "—", "nan", "#n/a", "#null!"}
)

_TRUE_TOKENS = frozenset({"true", "t", "yes", "y"})
_FALSE_TOKENS = frozenset({"false", "f", "no", "n"})

# Symbols stripped before a column is read as a number, and only when *every*
# non-blank value in that column carries the same shape. Deliberately not a general
# "remove punctuation": a comma is a thousands separator in one locale and a decimal
# point in another, so a column mixing "1,5" and "1,500" is left as text.
_CURRENCY = "$£€₹¥"

_INT_RE = re.compile(r"^[+-]?\d+$")
_FLOAT_RE = re.compile(r"^[+-]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?$")
_THOUSANDS_RE = re.compile(r"^[+-]?\d{1,3}(,\d{3})+(\.\d+)?$")

# Ordered: the first format that parses every value in the column wins, so an
# unambiguous ISO column is never re-read as something else.
_DATE_FORMATS: tuple[str, ...] = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%m-%d-%Y",
    "%d.%m.%Y",
    "%d %b %Y",
    "%d-%b-%Y",
    "%b %d %Y",
    "%d %B %Y",
)

_TIMESTAMP_FORMATS: tuple[str, ...] = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%d/%m/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%m/%d/%Y %H:%M",
)

# The two orders that cannot be told apart from the digits alone.
_DAY_FIRST = {"%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"}
_MONTH_FIRST = {"%m/%d/%Y", "%m-%d-%Y", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M"}

# SQLite is dynamically typed, but the *declared* type is what SQLAlchemy reflects
# and therefore what `connectors.sql.classify_type_family` classifies. Declaring
# DATE rather than TEXT is what makes an uploaded date column eligible as a KPI time
# field, so these names are load-bearing, not decoration.
_SQL_TYPES = {
    "INTEGER": "INTEGER",
    "REAL": "REAL",
    "DATE": "DATE",
    "TIMESTAMP": "TIMESTAMP",
    "BOOLEAN": "BOOLEAN",
    "TEXT": "TEXT",
}

_FAMILIES = {
    "INTEGER": "NUMERIC",
    "REAL": "NUMERIC",
    "DATE": "TEMPORAL",
    "TIMESTAMP": "TEMPORAL",
    "BOOLEAN": "BOOLEAN",
    "TEXT": "TEXT",
}


# ---------------------------------------------------------------------------
# What the caller gets back
# ---------------------------------------------------------------------------
@dataclass(slots=True)
class ColumnPlan:
    """One column as it will be created, and what was assumed to get there."""

    name: str
    header: str
    sql_type: str
    type_family: str
    filled: int = 0
    blank: int = 0
    # Values that matched a blank token rather than being genuinely empty. Kept
    # apart from `blank` because "this cell was empty" and "this cell said n/a" are
    # different facts about the export, and only the second one was interpreted.
    blank_tokens: int = 0
    distinct_sample: list[Any] = field(default_factory=list)
    # Every assumption made about this column, in the reader's words. Empty means
    # the values were taken exactly as written.
    notes: list[str] = field(default_factory=list)

    @property
    def blank_pct(self) -> float:
        total = self.filled + self.blank
        return round(self.blank / total * 100, 2) if total else 0.0


@dataclass(slots=True)
class SheetPlan:
    """One sheet, as one table."""

    table_name: str
    source_name: str
    columns: list[ColumnPlan]
    row_count: int
    skipped_rows: int = 0
    # A handful of whole rows, for the reader to recognise their own data before it
    # becomes a governed source.
    sample_rows: list[dict[str, Any]] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    # Held for the load step; not part of what the API returns.
    rows: list[list[Any]] = field(default_factory=list, repr=False)


@dataclass(slots=True)
class WorkbookPlan:
    filename: str
    file_format: str
    sheets: list[SheetPlan]
    notes: list[str] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(sheet.row_count for sheet in self.sheets)


# ---------------------------------------------------------------------------
# Names
# ---------------------------------------------------------------------------
def safe_identifier(raw: str, *, fallback: str, used: set[str]) -> str:
    """A header turned into something ``validate_identifier`` accepts.

    Lossy by necessity -- "Net Revenue (₹)" cannot be a SQL identifier -- so the
    original is always kept beside the result as the column comment. Collisions get
    a numeric suffix rather than silently overwriting, because two columns
    normalising to the same name is common ("Q1 %" and "Q1 (%)") and losing one
    would drop data without saying so.
    """

    cleaned = re.sub(r"[^0-9a-zA-Z]+", "_", (raw or "").strip()).strip("_").lower()
    cleaned = re.sub(r"_{2,}", "_", cleaned)
    if not cleaned:
        cleaned = fallback
    if cleaned[0].isdigit():
        cleaned = f"c_{cleaned}"
    cleaned = cleaned[:60]

    candidate = cleaned
    suffix = 2
    while candidate in used:
        candidate = f"{cleaned[:56]}_{suffix}"
        suffix += 1
    used.add(candidate)
    return validate_identifier(candidate, kind="column name")


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def _decode(data: bytes) -> tuple[str, str]:
    """Text plus the encoding that produced it.

    UTF-8 with BOM first because that is what Excel writes; cp1252 next because
    that is what everything else on Windows writes. The last fallback replaces
    undecodable bytes rather than failing, and says so -- a report naming three
    mangled characters is more useful than a refusal to open the file at all.
    """

    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "utf-8 (with replacements)"


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        # Sniffer fails on a single-column file, which is legitimate. Counting
        # candidates on the header line is a better guess than refusing.
        header = sample.splitlines()[0] if sample.splitlines() else ""
        counts = {sep: header.count(sep) for sep in (",", ";", "\t", "|")}
        best = max(counts, key=lambda key: counts[key])
        return best if counts[best] else ","


#: One parsed sheet: ``(name, header, rows, rows_skipped)``. The skipped count is
#: carried per sheet rather than as a file-wide total, because "41 rows were dropped"
#: is only actionable once you know which sheet to go and look at.
ParsedSheet = tuple[str, list[str], list[list[Any]], int]


def _read_delimited(data: bytes, *, max_rows: int) -> tuple[list[ParsedSheet], list[str]]:
    text, encoding = _decode(data)
    notes: list[str] = []
    if encoding != "utf-8-sig" and encoding != "utf-8":
        notes.append(f"Read as {encoding}.")

    delimiter = _sniff_delimiter(text[:16_384])
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)

    header: list[str] | None = None
    rows: list[list[Any]] = []
    skipped = 0
    for raw in reader:
        if header is None:
            if not any(str(cell).strip() for cell in raw):
                continue  # leading blank line before the header
            header = [str(cell) for cell in raw]
            continue
        if not any(str(cell).strip() for cell in raw):
            continue  # a blank separator line is not a row of data
        if len(raw) != len(header):
            skipped += 1
            continue
        rows.append(list(raw))
        if len(rows) >= max_rows:
            notes.append(f"Stopped after {max_rows:,} rows; the rest of the file was not read.")
            break

    if header is None:
        raise ValidationFailure("The file has no header row, so its columns cannot be named.")
    if delimiter != ",":
        notes.append(f"Columns separated by {'tab' if delimiter == chr(9) else delimiter!r}.")
    if skipped:
        notes.append(
            f"{skipped:,} row(s) had a different number of fields than the header and were "
            "skipped rather than padded."
        )

    name = "data"
    return [(name, header, rows, skipped)], notes


def _read_excel(data: bytes, *, max_rows: int) -> tuple[list[ParsedSheet], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ValidationFailure(
            "Excel files need the openpyxl package on the server. Save the sheet as CSV "
            "and upload that instead, or ask an administrator to install it."
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationFailure(
            "That file could not be opened as an Excel workbook. If it is an older .xls "
            "file, save it as .xlsx or .csv first."
        ) from exc

    sheets: list[ParsedSheet] = []
    notes: list[str] = []

    try:
        for worksheet in workbook.worksheets:
            header: list[str] | None = None
            rows: list[list[Any]] = []
            skipped = 0
            for raw in worksheet.iter_rows(values_only=True):
                values = list(raw)
                if header is None:
                    if not any(str(cell).strip() for cell in values if cell is not None):
                        continue
                    header = [("" if cell is None else str(cell)) for cell in values]
                    # Trailing empty header cells are Excel's formatting artefacts,
                    # not columns. Kept if anything below them has data.
                    while header and not header[-1].strip():
                        header.pop()
                    continue
                if not any(cell is not None and str(cell).strip() for cell in values):
                    continue
                if len(values) > len(header):
                    # Data beyond the last named column: unnamed, so it cannot be
                    # governed. Dropped and counted, never given a synthetic header.
                    if any(cell is not None and str(cell).strip() for cell in values[len(header) :]):
                        skipped += 1
                        continue
                    values = values[: len(header)]
                if len(values) < len(header):
                    values = values + [None] * (len(header) - len(values))
                rows.append(values)
                if len(rows) >= max_rows:
                    notes.append(
                        f"'{worksheet.title}' stopped after {max_rows:,} rows; the rest was not read."
                    )
                    break

            if header is None or not rows:
                # An empty sheet is normal in a real workbook and is not an error.
                notes.append(f"Sheet '{worksheet.title}' held no data rows and was skipped.")
                continue
            if skipped:
                notes.append(
                    f"'{worksheet.title}': {skipped:,} row(s) had values beyond the last named "
                    "column and were skipped rather than given invented headers."
                )
            sheets.append((worksheet.title, header, rows, skipped))
    finally:
        workbook.close()

    if not sheets:
        raise ValidationFailure("The workbook contains no sheet with a header row and data.")
    if len(sheets) > 1:
        notes.append(
            f"{len(sheets)} sheets were read as {len(sheets)} tables, which can be joined "
            "because they live in one source."
        )
    return sheets, notes


# ---------------------------------------------------------------------------
# Type inference
# ---------------------------------------------------------------------------
def _is_blank(value: Any) -> tuple[bool, bool]:
    """``(blank, matched_a_blank_token)``."""
    if value is None:
        return True, False
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return True, False
        if stripped.lower() in _BLANK_TOKENS:
            return True, True
    return False, False


def _numeric_from_text(values: list[str]) -> tuple[str | None, str | None]:
    """``(sql_type, note)`` if every value reads as one kind of number."""

    if all(_INT_RE.match(v) for v in values):
        # A leading zero is significant in a code -- "007" is a store, not seven --
        # so a column carrying one stays text even though it parses as an integer.
        if any(len(v.lstrip("+-")) > 1 and v.lstrip("+-").startswith("0") for v in values):
            return None, None
        if all(abs(int(v)) < 2**63 for v in values):
            return "INTEGER", None
    if all(_FLOAT_RE.match(v) for v in values):
        return "REAL", None

    # Only now consider stripping symbols, and only when the whole column agrees.
    stripped: list[str] = []
    saw_currency = False
    saw_thousands = False
    saw_percent = False
    for value in values:
        candidate = value.strip()
        # Parentheses come off first: the accounting convention wraps the whole
        # figure including its currency symbol, so "($120.25)" is a negative
        # $120.25 and stripping the symbol first would never find one.
        negated = False
        if candidate.startswith("(") and candidate.endswith(")"):
            candidate, negated = candidate[1:-1].strip(), True
        if candidate and candidate[0] in _CURRENCY:
            candidate, saw_currency = candidate[1:].strip(), True
        elif candidate and candidate[-1] in _CURRENCY:
            candidate, saw_currency = candidate[:-1].strip(), True
        if candidate.endswith("%"):
            candidate, saw_percent = candidate[:-1].strip(), True
        if _THOUSANDS_RE.match(candidate):
            candidate, saw_thousands = candidate.replace(",", ""), True
        if negated:
            candidate = f"-{candidate}"
        stripped.append(candidate)

    if not (saw_currency or saw_thousands or saw_percent):
        return None, None
    if not all(_INT_RE.match(v) or _FLOAT_RE.match(v) for v in stripped):
        return None, None

    removed = [
        label
        for label, seen in (
            ("a currency symbol", saw_currency),
            ("thousands separators", saw_thousands),
            ("a percent sign", saw_percent),
        )
        if seen
    ]
    note = (
        f"Read as a number after removing {', '.join(removed)}. "
        "The figures are unchanged; only the formatting was dropped."
    )
    if saw_percent:
        note += " Values are stored as written, so 12.5% is stored as 12.5, not 0.125."
    sql_type = "INTEGER" if all(_INT_RE.match(v) for v in stripped) else "REAL"
    return sql_type, note


def _temporal_from_text(values: list[str]) -> tuple[str | None, str | None, str | None]:
    """``(sql_type, format, note)`` if one format parses every value."""

    for formats, sql_type in ((_TIMESTAMP_FORMATS, "TIMESTAMP"), (_DATE_FORMATS, "DATE")):
        matching = [
            fmt
            for fmt in formats
            if all(_try_strptime(value, fmt) is not None for value in values)
        ]
        if not matching:
            continue

        chosen = matching[0]
        note = None
        # Day-first and month-first are indistinguishable when every day is 12 or
        # less. Rather than refuse the column -- which would cost the reader their
        # time field, and with it every date-based feature -- the assumption is made
        # and stated, so it can be checked and the file re-exported as ISO.
        ambiguous = any(fmt in _DAY_FIRST for fmt in matching) and any(
            fmt in _MONTH_FIRST for fmt in matching
        )
        if ambiguous:
            note = (
                f"Dates are written as {chosen.replace('%d', 'DD').replace('%m', 'MM').replace('%Y', 'YYYY')} "
                "and could be read either way round, so day-before-month was assumed. "
                "Re-upload with YYYY-MM-DD dates if that is wrong."
            )
        elif chosen not in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            note = f"Dates were read as {chosen} and stored in ISO form."
        return sql_type, chosen, note

    return None, None, None


def _try_strptime(value: str, fmt: str) -> datetime | None:
    try:
        return datetime.strptime(value.strip(), fmt)
    except (ValueError, TypeError):
        return None


def _classify(values: list[Any]) -> tuple[str, str | None, list[str]]:
    """``(sql_type, temporal_format, notes)`` for one column's non-blank values.

    Excel hands back values already typed; a CSV hands back strings. Both arrive
    here, and the native types are trusted before any string parsing is attempted --
    re-deriving a type from ``str(cell)`` would be a second guess at something the
    file already stated.
    """

    if not values:
        # Nothing to go on. Text is the only honest answer: any narrower type would
        # be a claim about data that is not there.
        return "TEXT", None, ["Every value in this column was blank, so it was created as text."]

    if all(isinstance(v, bool) for v in values):
        return "BOOLEAN", None, []
    if all(isinstance(v, datetime) for v in values):
        # A datetime whose clock is always midnight is a date that Excel stored with
        # a time component. Declaring DATE keeps it usable as a daily time field.
        if all(v.time() == time.min for v in values):
            return "DATE", None, []
        return "TIMESTAMP", None, []
    if all(isinstance(v, date) and not isinstance(v, datetime) for v in values):
        return "DATE", None, []
    if all(isinstance(v, int) and not isinstance(v, bool) for v in values):
        return "INTEGER", None, []
    if all(isinstance(v, int | float) and not isinstance(v, bool) for v in values):
        return "REAL", None, []

    text = [str(v).strip() for v in values]
    lowered = {v.lower() for v in text}
    if lowered <= (_TRUE_TOKENS | _FALSE_TOKENS):
        return "BOOLEAN", None, ["Read as true/false from the words in the column."]

    sql_type, note = _numeric_from_text(text)
    if sql_type:
        return sql_type, None, [note] if note else []

    sql_type, fmt, note = _temporal_from_text(text)
    if sql_type:
        return sql_type, fmt, [note] if note else []

    return "TEXT", None, []


# ---------------------------------------------------------------------------
# Value conversion
# ---------------------------------------------------------------------------
def _convert(value: Any, sql_type: str, fmt: str | None) -> Any:
    """One cell, as it will be stored. ``None`` for anything blank."""

    blank, _ = _is_blank(value)
    if blank:
        return None

    if sql_type == "BOOLEAN":
        if isinstance(value, bool):
            return int(value)
        return 1 if str(value).strip().lower() in _TRUE_TOKENS else 0

    if sql_type in ("DATE", "TIMESTAMP"):
        if isinstance(value, datetime):
            moment = value
        elif isinstance(value, date):
            moment = datetime.combine(value, time.min)
        else:
            moment = _try_strptime(str(value), fmt) if fmt else None
            if moment is None:
                return str(value).strip()
        # ISO, so SQLite's own date functions and SQLAlchemy's reflected DATE type
        # both read it back correctly. Any other form would compare as text.
        return moment.date().isoformat() if sql_type == "DATE" else moment.isoformat(sep=" ")

    if sql_type in ("INTEGER", "REAL"):
        if isinstance(value, int | float) and not isinstance(value, bool):
            return int(value) if sql_type == "INTEGER" else float(value)
        candidate = str(value).strip()
        # Same order as inference: unwrap the accounting negative, then the symbol.
        negated = candidate.startswith("(") and candidate.endswith(")")
        if negated:
            candidate = candidate[1:-1].strip()
        if candidate and candidate[0] in _CURRENCY:
            candidate = candidate[1:].strip()
        elif candidate and candidate[-1] in _CURRENCY:
            candidate = candidate[:-1].strip()
        if candidate.endswith("%"):
            candidate = candidate[:-1].strip()
        candidate = candidate.replace(",", "")
        if negated:
            candidate = f"-{candidate}"
        try:
            return int(candidate) if sql_type == "INTEGER" else float(candidate)
        except ValueError:
            # Should be unreachable: inference only chose a numeric type because
            # every value parsed. Returning None rather than raising means one
            # surprising cell cannot fail an entire load, and a null is visible in
            # the profile whereas a substituted zero would not be.
            return None

    return str(value).strip()


# ---------------------------------------------------------------------------
# The plan
# ---------------------------------------------------------------------------
def plan_workbook(filename: str, data: bytes, *, max_rows: int, sample_rows: int = 5) -> WorkbookPlan:
    """Parse a file and decide what tables and columns it becomes.

    Reads and infers but writes nothing, so a reader can be shown exactly what the
    platform is about to create -- including everything it had to assume -- and
    refuse it before an uploaded file becomes a governed source underneath a KPI.
    """

    if not data:
        raise ValidationFailure("The uploaded file is empty.")

    suffix = Path(filename or "").suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        raw_sheets, notes = _read_excel(data, max_rows=max_rows)
        file_format = "Excel workbook"
    elif suffix == ".xls":
        raise ValidationFailure(
            "The older .xls format cannot be read. Open it and save as .xlsx or .csv."
        )
    elif suffix in (".csv", ".tsv", ".txt", ""):
        raw_sheets, notes = _read_delimited(data, max_rows=max_rows)
        file_format = "Delimited text"
    else:
        raise ValidationFailure(
            f"Files of type '{suffix}' cannot be read. Upload a .csv, .tsv or .xlsx file."
        )

    used_tables: set[str] = set()
    sheets: list[SheetPlan] = []
    stem = Path(filename or "upload").stem

    for sheet_name, header, rows, skipped in raw_sheets:
        table_name = safe_identifier(
            sheet_name if len(raw_sheets) > 1 else (stem or sheet_name),
            fallback="uploaded_data",
            used=used_tables,
        )

        used_columns: set[str] = set()
        columns: list[ColumnPlan] = []
        formats: list[str | None] = []

        for index, raw_header in enumerate(header):
            column_values = [row[index] if index < len(row) else None for row in rows]
            present: list[Any] = []
            blank = 0
            tokens = 0
            for value in column_values:
                is_blank, was_token = _is_blank(value)
                if is_blank:
                    blank += 1
                    tokens += 1 if was_token else 0
                else:
                    present.append(value)

            sql_type, fmt, inference_notes = _classify(present)
            formats.append(fmt)

            column_notes = list(inference_notes)
            name = safe_identifier(raw_header, fallback=f"column_{index + 1}", used=used_columns)
            if name != (raw_header or "").strip().lower():
                column_notes.append(f'Original heading: "{raw_header}"')
            if tokens:
                column_notes.append(
                    f"{tokens:,} value(s) such as 'n/a' were treated as missing rather than as text."
                )

            seen: list[Any] = []
            for value in present:
                converted = _convert(value, sql_type, fmt)
                if converted is not None and converted not in seen:
                    seen.append(converted)
                if len(seen) >= 5:
                    break

            columns.append(
                ColumnPlan(
                    name=name,
                    header=raw_header,
                    sql_type=_SQL_TYPES[sql_type],
                    type_family=_FAMILIES[sql_type],
                    filled=len(present),
                    blank=blank,
                    blank_tokens=tokens,
                    distinct_sample=seen,
                    notes=column_notes,
                )
            )

        converted_rows = [
            [_convert(row[i] if i < len(row) else None, columns[i].sql_type, formats[i]) for i in range(len(columns))]
            for row in rows
        ]

        sheet_notes: list[str] = []
        if not any(column.type_family == "TEMPORAL" for column in columns):
            # Not an error -- a dimension table has no dates -- but it decides whether
            # this table can carry a date-based KPI, so it is said plainly here rather
            # than discovered later when a KPI refuses to save.
            sheet_notes.append(
                "No date column was recognised. A KPI measured over time needs one, so "
                "check the date format if you expected this table to have it."
            )

        sheets.append(
            SheetPlan(
                table_name=table_name,
                source_name=sheet_name,
                columns=columns,
                row_count=len(converted_rows),
                skipped_rows=skipped,
                sample_rows=[
                    {columns[i].name: row[i] for i in range(len(columns))}
                    for row in converted_rows[:sample_rows]
                ],
                notes=sheet_notes,
                rows=converted_rows,
            )
        )

    return WorkbookPlan(filename=filename or "upload", file_format=file_format, sheets=sheets, notes=notes)


# ---------------------------------------------------------------------------
# The load
# ---------------------------------------------------------------------------
@dataclass(slots=True)
class LoadReport:
    path: str
    tables: list[dict[str, Any]]
    total_rows: int


def load_into_sqlite(plan: WorkbookPlan, db_path: Path) -> LoadReport:
    """Write the planned tables into a SQLite file, replacing any of the same name.

    Replacing rather than appending is the honest behaviour for a snapshot source: a
    re-uploaded export is a *new* statement of the same facts, and appending would
    double every figure it contains. The old rows go, but the record of every load
    stays in the audit trail with its own checksum and row count, so the history of
    what was loaded when is not lost.

    Written with the stdlib driver rather than SQLAlchemy because this is the one
    place in the platform that *writes* to a source. Keeping it out of the connector
    layer is deliberate: ``SqlConnector`` refuses anything but a single SELECT, and
    that guarantee should not be loosened to accommodate uploads.
    """

    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path)
    tables: list[dict[str, Any]] = []
    try:
        for sheet in plan.sheets:
            columns_ddl = ", ".join(
                f'"{column.name}" {column.sql_type}' for column in sheet.columns
            )
            connection.execute(f'DROP TABLE IF EXISTS "{sheet.table_name}"')
            connection.execute(f'CREATE TABLE "{sheet.table_name}" ({columns_ddl})')
            if sheet.rows:
                placeholders = ", ".join("?" for _ in sheet.columns)
                connection.executemany(
                    f'INSERT INTO "{sheet.table_name}" VALUES ({placeholders})', sheet.rows
                )
            tables.append(
                {
                    "table_name": sheet.table_name,
                    "source_name": sheet.source_name,
                    "rows": sheet.row_count,
                    "columns": len(sheet.columns),
                }
            )
        connection.commit()
    finally:
        connection.close()

    return LoadReport(
        path=str(db_path), tables=tables, total_rows=sum(item["rows"] for item in tables)
    )
